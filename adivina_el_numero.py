import random
import time
import openpyxl
import os
from getpass import getpass


def menu_princ():
    print("\n1. Partida modo solitario")
    print("2. Partida 2 Jugadores")
    print("3. Estadística")
    print("4. Salir\n")
    
    
def menu_dif():
    print("1. Fácil   (20 intentos)")
    print("2. Medio   (12 intentos)")
    print("3. Difícil (5 intentos)\n")
    
    
def juego_facil(numero):
    print("\n¡A JUGAR!")
    oportunidades = 0

    while oportunidades < 20:
        try:
            intento = int(input("Elige un número: "))
            if intento < numero:
                print("Debes buscar un número mayor")
            elif intento > numero:
                print("Debes buscar un número menor")
            else:
                print("\n-- ¡¡¡CORRECTO!!! --\n-- Enhorabuena, ¡has ganado! --\n")
                nombre_jugador = input("Por favor, ingresa tu nombre: ")
                guardar_resultado(nombre_jugador, "GANADO", "FÁCIL")
                break
            oportunidades += 1
        
        except ValueError:
            print("Tienes que introducir un número para poder jugar")        

    if oportunidades == 20:
        print("\n-- GAME OVER ¡Seguro que a la próxima lo consigues! --")
        print("-- El número buscado era", numero, " --\n")
        nombre_jugador = input("Por favor, ingresa tu nombre: ")
        guardar_resultado(nombre_jugador, "PERDIDO", "FÁCIL")
    

                      
def juego_medio(numero):
    print("\n¡A JUGAR!")
    oportunidades = 0

    while oportunidades < 12:
        try:
            intento = int(input("Elige un número: "))
            if intento < numero:
                print("Debes buscar un número mayor")
            elif intento > numero:
                print("Debes buscar un número menor")
            else:
                print("\n-- ¡¡¡CORRECTO!!! --\n-- Enhorabuena, ¡has ganado! --\n")
                nombre_jugador = input("Por favor, ingresa tu nombre: ")
                guardar_resultado(nombre_jugador, "GANADO", "MEDIA")
                break
            oportunidades += 1
        
        except ValueError:
            print("Tienes que introducir un número para poder jugar")        

    if oportunidades == 12:
        print("\n-- GAME OVER ¡Seguro que a la próxima lo consigues! --")
        print("-- El número buscado era", numero, " --\n")
        nombre_jugador = input("Por favor, ingresa tu nombre: ")
        guardar_resultado(nombre_jugador, "PERDIDO", "MEDIA")
    
                 
def juego_dificil(numero):
    print("\n¡A JUGAR!")
    oportunidades = 0

    while oportunidades < 5:
        try:
            intento = int(input("Elige un número: "))
            if intento < numero:
                print("Debes buscar un número mayor")
            elif intento > numero:
                print("Debes buscar un número menor")
            else:
                print("\n-- ¡¡¡CORRECTO!!! --\n-- Enhorabuena, ¡has ganado! --\n")
                nombre_jugador = input("Por favor, ingresa tu nombre: ")
                guardar_resultado(nombre_jugador, "GANADO", "DIFÍCIL")
                break
            oportunidades += 1
        
        except ValueError:
            print("Tienes que introducir un número para poder jugar")        

    if oportunidades == 5:
        print("\n-- GAME OVER ¡Seguro que a la próxima lo consigues! --")
        print("-- El número buscado era", numero, " --\n")
        nombre_jugador = input("Por favor, ingresa tu nombre: ")
        guardar_resultado(nombre_jugador, "PERDIDO", "DIFÍCIL")
        
        
def guardar_resultado(nombre_jugador, resultado, nivel_dificultad):
    archivo_excel = 'estadisticas_juego_python.xlsx'

    if os.path.isfile(archivo_excel):
        libro = openpyxl.load_workbook(archivo_excel)
    else:
        libro = openpyxl.Workbook()
        
    hoja = libro.active
    hoja.append([nombre_jugador, resultado, nivel_dificultad])
    libro.save(archivo_excel)


#COMIENZO DEL CÓDIGO DEL JUEGO   

print('BIENVENIDO A "ADIVIDA EL NÚMERO". ¡DISFRUTA DEL JUEGO!')
print("Las reglas son claras, debes adivinar un número entre 1 y 1000... ambos incluidos. ¡Suerte!")
while True:
    menu_princ()    
    menu_elec = input("Elige una opción --> ")

    while menu_elec not in ["1", "2", "3", "4"]:
        print("Por favor, selecciona una opción del menú utilizando los números del 1 al 4")
        menu_elec = input("¡Elige una opción! ")
    else:
        if menu_elec == "1":
            print("\nHas seleccionado partida de un jugador\n\nPuedes elegir entre 3 dificultades\n")
            menu_dif()

            dificultad = input("Elige una dificultad ")

            while dificultad not in ["1", "2", "3"]:
                print("Por favor, selecciona una dificultad válida")
                dificultad = input("Elige una dificultad --> ")

            else:    
                if dificultad == "1":
                    print("\n¡Has seleccionado dificultad fácil! Tienes 20 intentos para adivinar el número")
                    numero = random.randint(1, 1000)
                    juego_facil(numero)

                elif dificultad == "2":
                    print("\n¡Has seleccionado dificultad media! Tienes 12 intentos para adivinar el número")
                    numero = random.randint(1, 1000)
                    juego_medio(numero)

                elif dificultad == "3":
                    print("\n¡Has seleccionado dificultad difícil! Tienes 5 intentos para adivinar el número")
                    numero = random.randint(1, 1000)
                    juego_dificil(numero)


        elif menu_elec == "2":
            print("\nHas seleccionado partida de dos jugadores\n\nPuedes elegir entre 3 dificultades\n")
            menu_dif()

            dificultad = input("Elige una dificultad ")

            while dificultad not in ["1", "2", "3"]:
                print("Por favor, selecciona una dificultad válida")
                dificultad = input("Elige una dificultad ")

            else:
                if dificultad == "1":
                    print("\nHas seleccionado dificultad fácil! Tienes 20 intentos para adivinar el número")
                elif dificultad == "2":
                    print("\nHas seleccionado dificultad media! Tienes 12 intentos para adivinar el número")
                elif dificultad == "3":
                    print("\nHas seleccionado dificultad difícil! Tienes 5 intentos para adivinar el número")

                while True:
                    try:
                        numero = int(getpass("Ingresa el número a adivinar entre 1 y 1000: "))
                        if 1 <= numero <= 1000:
                            break
                        else:
                            print("El número debe encontrarse entre 1 y 1000")
                    except ValueError:
                        print("¡Recuerda! Debes ingresar un número")

                if dificultad == "1":
                    juego_facil(numero)
                elif dificultad == "2":
                    juego_medio(numero)
                elif dificultad == "3":
                    juego_dificil(numero)


        elif menu_elec == "3":
            estadisticas_excel = 'estadisticas_juego_python.xlsx'

            if os.path.isfile(estadisticas_excel):
                excel = openpyxl.load_workbook(estadisticas_excel)
                hoja = excel.active
                print("\nEstadísticas:")
                for fila in hoja.iter_rows(min_row=1, max_col=3, values_only=True):
                    nombre_jugador, resultado, nivel_dificultad = fila
                    print(nombre_jugador, "ha", resultado, "la partida en dificultad", nivel_dificultad)
            else:
                print("\n¡Vaya! Parece que aún no has jugado ninguna partida por lo que no hay nada que mostrar :(")

        elif menu_elec == "4":
            print("\nCerrando juego\n")
            time.sleep(0.5)
            print("·")
            time.sleep(0.5)
            print("·")
            time.sleep(0.5)
            print("·")
            time.sleep(0.5)
            break
exit()
